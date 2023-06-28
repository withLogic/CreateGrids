from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins, PrintOptions, PrintPageSetup
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import datetime
from datetime import timedelta
from pandas import *

import math

def smart_truncate(content, length=35, suffix=''):
    if len(content) <= length:
        return content
    else:
        return ' '.join(content[:length+1].split(' ')[0:-1]) + suffix

def create_workbook(path):
    data = read_csv("a2z_report.csv")

    dayList = data["Session Start Date"].tolist();
    dayList = sorted(list(set(dayList)))
    dayList.reverse();

    timeList = data["Session Start Time"].tolist();
    timeList = list(set(timeList))
    timeList.sort(key=lambda date: datetime.datetime.strptime(date, "%I:%M %p"))
    timeList.reverse()

    roomList = data["Room Number"].tolist();
    roomList = sorted(list(set(roomList)))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = dayList.pop().replace("/","-")

    while dayList:
    	workbook.create_sheet(dayList.pop().replace("/","-"))

    startTime = datetime.datetime.strptime(timeList.pop(), "%I:%M %p")

    ft = Font(name="Arial Narrow", size="10", color="000000", bold=False)
    fth = Font(name="Arial Narrow", size="12", color="000000", bold=True)
    ag = Alignment(horizontal="center", vertical="center")
    agc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdlr = Border(left=Side(border_style='thick', color='00000000'),
					right=Side(border_style='thick', color='00000000'),
					top=Side(border_style=None, color='FF000000'),
					bottom=Side(border_style=None, color='FF000000'),
					diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
					outline=Side(border_style=None, color='FF000000'),
					vertical=Side(border_style=None, color='FF000000'),
					horizontal=Side(border_style=None, color='FF000000')
				)
    bdtlr = Border(left=Side(border_style='thick', color='00000000'),
					right=Side(border_style='thick', color='00000000'),
					top=Side(border_style='thick', color='FF000000'),
					bottom=Side(border_style=None, color='FF000000'),
					diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
					outline=Side(border_style=None, color='FF000000'),
					vertical=Side(border_style=None, color='FF000000'),
					horizontal=Side(border_style=None, color='FF000000')
				)
    bdblr = Border(left=Side(border_style='thick', color='00000000'),
					right=Side(border_style='thick', color='00000000'),
					top=Side(border_style=None, color='FF000000'),
					bottom=Side(border_style='thick', color='FF000000'),
					diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
					outline=Side(border_style=None, color='FF000000'),
					vertical=Side(border_style=None, color='FF000000'),
					horizontal=Side(border_style=None, color='FF000000')
				)
    bdb = Border(left=Side(border_style=None, color='00000000'),
					right=Side(border_style=None, color='00000000'),
					top=Side(border_style=None, color='FF000000'),
					bottom=Side(border_style='thick', color='FF000000'),
					diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
					outline=Side(border_style=None, color='FF000000'),
					vertical=Side(border_style=None, color='FF000000'),
					horizontal=Side(border_style=None, color='FF000000')
				)
    bda = Border(left=Side(border_style='thick', color='00000000'),
					right=Side(border_style='thick', color='00000000'),
					top=Side(border_style='thick', color='FF000000'),
					bottom=Side(border_style='thick', color='FF000000'),
					diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
					outline=Side(border_style=None, color='FF000000'),
					vertical=Side(border_style=None, color='FF000000'),
					horizontal=Side(border_style=None, color='FF000000')
				)

    # Iterate through the sheets
    for sheet in workbook.worksheets:

	    # Set some variables
	    rowi = 1
	    coli = 1
	    countOfRooms = len(roomList) + 1
	    nextTime = startTime
	    sheetTitleAsDate = datetime.datetime.strptime(sheet.title,  '%m-%d-%Y')

	    # set Widths
	    sheet.column_dimensions["A"].width = 4.78
	    sheet.column_dimensions[chr(65 + countOfRooms)].width = 4.78
	    sheet.row_dimensions[1].height = 15.75
	    for room in range(1, countOfRooms):
	    	sheet.column_dimensions[chr(65 + room)].width = 18.5

	    # set page margins
	    sheet.page_margins = PageMargins(left=0.25, right=0.25, top=1, bottom=1, header=0.5, footer=0.5)
	    sheet.print_options = PrintOptions(horizontalCentered=True, verticalCentered=True, headings=False, gridLines=True, gridLinesSet=True)
	    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
	    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
	    sheet.page_setup.fitToPage = True

	    # set the header and footer
	    sheet.oddHeader.center.text = "Annual Gathering 2023 - Baltimore, MD \n " + sheetTitleAsDate.strftime("%A, %B %m, %Y") + " \n Program Grid"
	    sheet.oddHeader.center.size = 20
	    sheet.oddHeader.center.font = "Calibri,Bold"
	    sheet.oddHeader.center.color = "000000"

	    sheet.oddFooter.center.text = "Shop the Mensa Store and wear your brain on your sleeve with our exclusive licensed apparel."
	    sheet.oddFooter.center.size = 20
	    sheet.oddFooter.center.font = "Calibri"
	    sheet.oddFooter.center.color = "000000"

	    # Create the dates in the upper corners
	    sheet.cell(row=rowi, column=coli).value = sheetTitleAsDate.strftime("%a")
	    sheet.cell(row=rowi, column=coli).font = fth
	    sheet.cell(row=rowi, column=coli).alignment = ag
	    sheet.cell(row=rowi, column=coli).border = bdlr
	    sheet.cell(row=rowi, column=coli + countOfRooms).value = sheetTitleAsDate.strftime("%a")
	    sheet.cell(row=rowi, column=coli + countOfRooms).font = fth
	    sheet.cell(row=rowi, column=coli + countOfRooms).alignment = ag
	    sheet.cell(row=rowi, column=coli + countOfRooms).border = bdlr
	    rowi = rowi + 1;

	    sheet.cell(row=rowi, column=coli).value = sheetTitleAsDate.strftime("%#m/%#d")
	    sheet.cell(row=rowi, column=coli).font = fth
	    sheet.cell(row=rowi, column=coli).alignment = ag
	    sheet.cell(row=rowi, column=coli).border = bdlr
	    sheet.cell(row=rowi, column=coli + countOfRooms).value = sheetTitleAsDate.strftime("%#m/%#d")
	    sheet.cell(row=rowi, column=coli + countOfRooms).font = fth
	    sheet.cell(row=rowi, column=coli + countOfRooms).alignment = ag
	    sheet.cell(row=rowi, column=coli + countOfRooms).border = bdlr
	    rowi = rowi + 1;

	    # Create the times on left and right.
	    while nextTime.time() != datetime.datetime.strptime('00:30:00',  '%H:%M:%S').time():
	    	sheet.cell(row=rowi, column=coli).value = nextTime.strftime('%#I:%M')
	    	sheet.cell(row=rowi, column=coli).font = fth
	    	sheet.cell(row=rowi, column=coli).alignment = ag

	    	sheet.cell(row=rowi, column=coli + countOfRooms).value = nextTime.strftime('%#I:%M')
	    	sheet.cell(row=rowi, column=coli + countOfRooms).font = fth
	    	sheet.cell(row=rowi, column=coli + countOfRooms).alignment = ag

	    	if(nextTime.time() == startTime.time()):
	    		sheet.cell(row=rowi, column=coli).border = bdtlr
	    		sheet.cell(row=rowi, column=coli + countOfRooms).border = bdtlr
	    	elif(nextTime.time() == datetime.datetime.strptime('00:00:00',  '%H:%M:%S').time()):
	    		sheet.cell(row=rowi, column=coli).border = bdblr
	    		sheet.cell(row=rowi, column=coli + countOfRooms).border = bdblr
	    	else:
	    		sheet.cell(row=rowi, column=coli).border = bdlr
	    		sheet.cell(row=rowi, column=coli + countOfRooms).border = bdlr

	    	nextTime = nextTime + timedelta(minutes=30)
	    	rowi = rowi + 1

	    # Create the room header
	    coli = coli + 1;
	    rowi = 1;
	    for roomIndex, roomName in enumerate(roomList):
	    	sheet.merge_cells(start_row=rowi, start_column=coli, end_row=rowi+1, end_column=coli)
	    	sheet.cell(row=rowi, column=coli).value = roomList[roomIndex]
	    	sheet.cell(row=rowi, column=coli).font = fth
	    	sheet.cell(row=rowi, column=coli).alignment = ag
	    	sheet.cell(row=rowi + 1, column=coli).border = bdb
	    	coli = coli + 1;

	    # Find the data for the day we're on
	    for roomIndex, roomName in enumerate(roomList):

		    dd = data.loc[(data['Session Start Date'] == sheet.title.replace("-","/")) & (data['Room Number'] == roomName)]

		    rowi = 3
		    coli = 2 + roomIndex
		    
		    if(len(dd)):
			    for ind in dd.index:
			    	sesStartTime = datetime.datetime.strptime(dd["Session Start Time"][ind], "%I:%M %p")
			    	sesEndTime = datetime.datetime.strptime(dd["Session End Time"][ind], "%I:%M %p")

			    	startInterval = ((sesStartTime - startTime).total_seconds() / 60) / 30
			    	endInterval = ((sesEndTime - sesStartTime).total_seconds() / 60) / 30

			    	rows = math.ceil((rowi + startInterval))
			    	rowe = math.ceil((rows + endInterval)) - 1

			    	sheet.cell(row=rows, column=coli).value = smart_truncate(dd["Session Title"][ind])
			    	sheet.cell(row=rows, column=coli).font = ft
			    	sheet.cell(row=rows, column=coli).alignment = agc
			    	sheet.cell(row=rows, column=coli).border = bda
			    	sheet.merge_cells(start_row=rows, start_column=coli, end_row=rowe, end_column=coli)

    workbook.save(path)

if __name__ == "__main__":
    create_workbook("a2z_report_out.xlsx")