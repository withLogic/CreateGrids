from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins, PrintOptions, PrintPageSetup
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import utils
from openpyxl.utils import get_column_letter

import datetime
from datetime import timedelta
from pandas import *

import math

import sys
from pathlib import Path

EVENT_NAME = "Annual Gathering 2025 - Chicago, IL"
START_DATE = datetime.datetime(2025, 7, 2)
ROOMS_TO_SUPPRESS = ['Ask for location', 'Private Dining Room 7']
HIDE_ROOMS_WITH_NO_SESSIONS = 1
INTERVAL = 15
REQUIRED_COLUMNS = [
    "Session Title",
    "Session Start Date",
    "Session Start Time",
    "Session End Time",
    "Room"
]
FOOTER_TEXT = "Shop the Mensa Store and wear your brain on your sleeve with our exclusive licensed apparel."

RED = "\033[91m"
YELLOW = "\033[93m"
BLUE = "\033[94m"
RESET = "\033[0m"

def smart_truncate(content, length=50, suffix=''):
    if len(content) <= length:
        return content
    else:
        return ' '.join(content[:length+1].split(' ')[0:-1]) + suffix

def create_workbook(path, inputfile, outputfile):

    print(f"{YELLOW}[INFO]{RESET} Creating Event Grid from {BLUE}{inputfile}{RESET}.")

    if inputfile.lower().endswith(".csv"):
        data = read_csv(path / inputfile)
    elif inputfile.lower().endswith(".xls"):
        data_raw = read_html(path / inputfile)
        table_raw = data_raw[0]
        table_header = table_raw.iloc[0]
        data = table_raw.iloc[1:].copy()

        data.columns = table_header
        data.rename(columns={"Room Number": "Room"}, inplace=True)
        data.reset_index(drop=True, inplace=True)
    else:
        print(f"{RED}[ERROR]{RESET} Unable to find a valid input file. Supported file types are CSV and XLS.")
        exit()

    columnsInData = [col.strip() for col in data.columns]
    missingColumns = [col for col in REQUIRED_COLUMNS if col not in columnsInData]

    if missingColumns:
        print(f"{RED}[ERROR]{RESET} Input file is missing required columns: {BLUE}{', '.join(missingColumns)}{RESET}")
        exit()

    dayList = data["Session Start Date"].tolist()
    dayList = sorted(list(set(dayList)))
    dayList.reverse()
    dayList = [x for x in dayList if datetime.datetime.strptime(x, '%m/%d/%Y') >= START_DATE]

    timeList = data["Session Start Time"].tolist()
    timeList = list(set(timeList))
    timeList.sort(key=lambda date: datetime.datetime.strptime(date, "%I:%M %p"))
    timeList.reverse()

    roomList = data["Room"].tolist()
    roomList = sorted(list(set(roomList)))
    roomList = [x for x in roomList if x not in ROOMS_TO_SUPPRESS]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = dayList.pop().replace("/","-")

    while dayList:
        workbook.create_sheet(dayList.pop().replace("/","-"))

    startTime = datetime.datetime.strptime(timeList.pop(), "%I:%M %p")

    ft = Font(name="Arial Narrow", size="10", color="000000", bold=False)
    fth = Font(name="Arial Narrow", size="12", color="000000", bold=True)
    ag = Alignment(horizontal="center", vertical="center")
    agc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdlr = Border(left=Side(border_style='thick', color='00000000'),
                    right=Side(border_style='thick', color='00000000'),
                    top=Side(border_style=None, color='FF000000'),
                    bottom=Side(border_style=None, color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                    outline=Side(border_style=None, color='FF000000'),
                    vertical=Side(border_style=None, color='FF000000'),
                    horizontal=Side(border_style=None, color='FF000000')
                )
    bdtlr = Border(left=Side(border_style='thick', color='00000000'),
                    right=Side(border_style='thick', color='00000000'),
                    top=Side(border_style='thick', color='FF000000'),
                    bottom=Side(border_style=None, color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                    outline=Side(border_style=None, color='FF000000'),
                    vertical=Side(border_style=None, color='FF000000'),
                    horizontal=Side(border_style=None, color='FF000000')
                )
    bdblr = Border(left=Side(border_style='thick', color='00000000'),
                    right=Side(border_style='thick', color='00000000'),
                    top=Side(border_style=None, color='FF000000'),
                    bottom=Side(border_style='thick', color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                    outline=Side(border_style=None, color='FF000000'),
                    vertical=Side(border_style=None, color='FF000000'),
                    horizontal=Side(border_style=None, color='FF000000')
                )
    bdb = Border(left=Side(border_style=None, color='00000000'),
                    right=Side(border_style=None, color='00000000'),
                    top=Side(border_style=None, color='FF000000'),
                    bottom=Side(border_style='thick', color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                    outline=Side(border_style=None, color='FF000000'),
                    vertical=Side(border_style=None, color='FF000000'),
                    horizontal=Side(border_style=None, color='FF000000')
                )
    bda = Border(left=Side(border_style='thick', color='00000000'),
                    right=Side(border_style='thick', color='00000000'),
                    top=Side(border_style='thick', color='FF000000'),
                    bottom=Side(border_style='thick', color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                    outline=Side(border_style=None, color='FF000000'),
                    vertical=Side(border_style=None, color='FF000000'),
                    horizontal=Side(border_style=None, color='FF000000')
                )

    # Iterate through the sheets
    for sheet in workbook.worksheets:

        # Set some variables
        rowi = 1
        coli = 1
        countOfRooms = len(roomList) + 1
        nextTime = startTime
        sheetTitleAsDate = datetime.datetime.strptime(sheet.title,  '%m-%d-%Y')

        # set Widths
        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions[utils.cell.get_column_letter(countOfRooms + 1)].width = 5
        sheet.row_dimensions[1].height = 15.75
        for room in range(1, countOfRooms):
            col_letter = get_column_letter(room + 1)
            sheet.column_dimensions[col_letter].width = 18.5

        # set page margins
        sheet.page_margins = PageMargins(left=0.25, right=0.25, top=1, bottom=1, header=0.5, footer=0.5)
        sheet.print_options = PrintOptions(horizontalCentered=True, verticalCentered=True, headings=False, gridLines=True, gridLinesSet=True)
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
        sheet.page_setup.fitToPage = True
        sheet.print_area = 'A1:' + utils.cell.get_column_letter(countOfRooms + 1) + '36'

        # set the header and footer
        sheet.oddHeader.center.text = EVENT_NAME + " \n " + sheetTitleAsDate.strftime("%A, %B %m, %Y") + " \n Program Grid"
        sheet.oddHeader.center.size = 20
        sheet.oddHeader.center.font = "Calibri,Bold"
        sheet.oddHeader.center.color = "000000"

        sheet.oddFooter.center.text = FOOTER_TEXT
        sheet.oddFooter.center.size = 20
        sheet.oddFooter.center.font = "Calibri"
        sheet.oddFooter.center.color = "000000"

        # Create the dates in the upper corners
        sheet.cell(row=rowi, column=coli).value = sheetTitleAsDate.strftime("%a")
        sheet.cell(row=rowi, column=coli).font = fth
        sheet.cell(row=rowi, column=coli).alignment = ag
        sheet.cell(row=rowi, column=coli).border = bdlr
        sheet.cell(row=rowi, column=coli + countOfRooms).value = sheetTitleAsDate.strftime("%a")
        sheet.cell(row=rowi, column=coli + countOfRooms).font = fth
        sheet.cell(row=rowi, column=coli + countOfRooms).alignment = ag
        sheet.cell(row=rowi, column=coli + countOfRooms).border = bdlr
        rowi = rowi + 1

        sheet.cell(row=rowi, column=coli).value = sheetTitleAsDate.strftime("%#m/%#d")
        sheet.cell(row=rowi, column=coli).font = fth
        sheet.cell(row=rowi, column=coli).alignment = ag
        sheet.cell(row=rowi, column=coli).border = bdlr
        sheet.cell(row=rowi, column=coli + countOfRooms).value = sheetTitleAsDate.strftime("%#m/%#d")
        sheet.cell(row=rowi, column=coli + countOfRooms).font = fth
        sheet.cell(row=rowi, column=coli + countOfRooms).alignment = ag
        sheet.cell(row=rowi, column=coli + countOfRooms).border = bdlr
        rowi = rowi + 1

        # Create the times on left and right.
        timeInvervalRowCount = rowi + int((24*60)/INTERVAL)
        while rowi != timeInvervalRowCount:
            sheet.cell(row=rowi, column=coli).value = nextTime.strftime('%#I:%M')
            sheet.cell(row=rowi, column=coli).font = fth
            sheet.cell(row=rowi, column=coli).alignment = ag

            sheet.cell(row=rowi, column=coli + countOfRooms).value = nextTime.strftime('%#I:%M')
            sheet.cell(row=rowi, column=coli + countOfRooms).font = fth
            sheet.cell(row=rowi, column=coli + countOfRooms).alignment = ag

            if(nextTime.time() == startTime.time()):
                sheet.cell(row=rowi, column=coli).border = bdtlr
                sheet.cell(row=rowi, column=coli + countOfRooms).border = bdtlr
            elif(nextTime.time() == datetime.datetime.strptime('00:00:00',  '%H:%M:%S').time()):
                sheet.cell(row=rowi, column=coli).border = bdblr
                sheet.cell(row=rowi, column=coli + countOfRooms).border = bdblr
            else:
                sheet.cell(row=rowi, column=coli).border = bdlr
                sheet.cell(row=rowi, column=coli + countOfRooms).border = bdlr

            nextTime = nextTime + timedelta(minutes=INTERVAL)
            rowi = rowi + 1

        # Create the room header
        coli = coli + 1
        rowi = 1
        for roomIndex, roomName in enumerate(roomList):
            sheet.merge_cells(start_row=rowi, start_column=coli, end_row=rowi+1, end_column=coli)
            sheet.cell(row=rowi, column=coli).value = roomList[roomIndex]
            sheet.cell(row=rowi, column=coli).font = fth
            sheet.cell(row=rowi, column=coli).alignment = agc
            sheet.cell(row=rowi + 1, column=coli).border = bdb

            coli = coli + 1

        currentSheetRooms = set(data.loc[(data['Session Start Date'] == sheet.title.replace("-","/"))]["Room"].dropna())
        emptyRooms = [(i, room) for i, room in enumerate(roomList) if room not in currentSheetRooms]

        # Find the data for the day we're on
        for roomIndex, roomName in enumerate(roomList):

            dd = data.loc[(data['Session Start Date'] == sheet.title.replace("-","/")) & (data['Room'] == roomName)]
            rowi = 3
            coli = 2 + roomIndex
            
            if(len(dd)):
                for ind in dd.index:
                    sesStartTime = datetime.datetime.strptime(dd["Session Start Time"][ind], "%I:%M %p")
                    sesEndTime = datetime.datetime.strptime(dd["Session End Time"][ind], "%I:%M %p")

                    startInterval = ((sesStartTime - startTime).total_seconds() / 60) / INTERVAL
                    endInterval = ((sesEndTime - sesStartTime).total_seconds() / 60) / INTERVAL

                    rows = math.ceil((rowi + startInterval))
                    rowe = math.ceil((rows + endInterval)) - 1

                    if "DO NOT PUBLISH" in dd["Session Title"][ind]:
                        print(f"{YELLOW}[INFO]{RESET} {BLUE}{dd["Session Title"][ind]}{RESET} is not set to publish. Skipping.")
                        continue

                    try:
                        sheet.cell(row=rows, column=coli).value = smart_truncate(dd["Session Title"][ind])
                        sheet.cell(row=rows, column=coli).font = ft
                        sheet.cell(row=rows, column=coli).alignment = agc
                        sheet.cell(row=rows, column=coli).border = bda
                        sheet.merge_cells(start_row=rows, start_column=coli, end_row=rowe, end_column=coli)
                    except AttributeError as e:
                        print(f"{RED}[ERROR]{RESET} Overlapping session schedule. {BLUE}{dd["Session Title"][ind]}{RESET} overaps with another session. Skipping.")
                        pass

        # hide the cells that we do not need
        if HIDE_ROOMS_WITH_NO_SESSIONS:
            for index, room in emptyRooms:
                col_letter = get_column_letter(index + 2)
                sheet.column_dimensions[col_letter].hidden = True

                print(f"{YELLOW}[INFO]{RESET} {BLUE}{room}{RESET} does not have any sessions on {BLUE}{sheet.title}{RESET}. Setting the column to Hidden.")

    workbook.save(path / outputfile)
    print(f"{YELLOW}[INFO]{RESET} Event Grid has been saved here {BLUE}{outputfile}{RESET}.")

schedgrid = sys.argv[-1]

if __name__ == "__main__":
    create_workbook(Path.cwd(), schedgrid, "event_grid__out.xlsx")